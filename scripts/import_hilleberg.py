import sys
import pandas as pd
import itertools
import openpyxl

import psycopg2

# Fetch environment variables
dbname = os.environ.get('DB_NAME', 'fbsr_pontun')
user = os.environ.get('DB_USER', 'fbsr_pontun')
host = os.environ.get('DB_HOST', '127.0.0.1')
password = os.environ.get('DB_PASSWORD', '')

# Use the environment variables in the connection string
conn = psycopg2.connect(f"dbname={dbname} user={user} host={host} password={password}")

cur  = conn.cursor()

def main(listing_id, filename):
    df = pd.read_excel(filename)
    workbook = openpyxl.load_workbook(filename, data_only=True)
    worksheet = workbook[' DROP 1']
    saved, skipped, added = 0, 0, set()
    category = None
    for row in range(11,worksheet.max_row+1):
        _row = []
        vendor_id = worksheet.cell(column=1, row=row).value
        model     = worksheet.cell(column=2, row=row).value
        color     = worksheet.cell(column=3, row=row).value
        price_excl_disc = worksheet.cell(column=5, row=row).value
        discount = worksheet.cell(column=8, row=row).value
        if vendor_id is None and model == "Nallo 3 GT":
            vendor_id = "0213361"
        vendor_id = str(vendor_id)
        if vendor_id.startswith("9"): continue
        elif not vendor_id.startswith("0"):
            category = vendor_id
            continue


        #price = price_excl_disc * (1 - discount)
        print(f"{vendor_id} {category} {model} {color} {price_excl_disc} {discount}")
        model = model.strip()
        if color is not None:
            color = color.strip()
        vendor_id = vendor_id.strip()

        if category in ('Yellow Label', 'Red Label', 'Black Label'):
            url = f'https://hilleberg.com/eng/tent/{category.lower().replace(" ", "-")}-tents/{model.lower().replace(" ", "-")}/'
        elif category == 'Shelters' and model == 'Bivanorak':
            url = 'https://hilleberg.com/eng/shelters/bivanorak/'
        elif category == 'Shelters' and model.startswith("Tarp"):
            url = 'https://hilleberg.com/eng/shelters/tarp/'
        elif category == 'Shelters' and model.startswith('Windsack'):
            url = 'https://hilleberg.com/eng/shelters/windsack/'
        elif category in ('Spare Poles', 'Pegs', 'Footprint', 'Pole Holder', 'Repair Materials', 'Stuff Sacks For Pegs, Poles And Tents & Other Bags', 'Pole Holder And Guy Lines'):
            url = 'https://hilleberg.com/eng/tent/tent-accessories/'
        elif category == 'Mesh Inner Tents':
            url = 'https://hilleberg.com/eng/shelters/mesh-inner-tents/'
        else:
            url = None

        if not pd.isnull(price_excl_disc):
            cur.execute(f"INSERT INTO items (listing_id, item_name, vendor_id, price, vendor_url) VALUES (%s, %s, %s, %s, %s) RETURNING id", (listing_id, f"{model} ({category})", vendor_id, price_excl_disc, url))
            item_id = cur.fetchone()[0]
            if color is not None:
                cur.execute(f"INSERT INTO item_properties (item_id, original, adjusted, value) VALUES (%s, %s, %s, %s)", (item_id, "Litur", "Litur", color))
        #    #print(f"{listing_id} : {name} : {vendor_id} : {price} : {desc} : {url}")
            saved += 1
        else:
            if not pd.isnull(vendor_id):
                print(f"Skipped row item {vendor_id} : {name}")
            skipped += 1
    #print(f"Saved {saved} skipped {skipped} pandas df has {df.loc[~pd.isnull(pd.to_numeric(df.iloc[:,4], errors='coerce'))].shape[0]} rows")

    conn.commit()
    import pdb; pdb.set_trace()



if __name__ == "__main__":
    assert len(sys.argv) > 2, "Missing listing id or filename"
    main(int(sys.argv[1]), sys.argv[2])
