import sys
import pandas as pd
import itertools
import openpyxl

import psycopg2

# Fetch environment variables
dbname = os.environ.get('DB_NAME', 'fbsr_pontun')
user = os.environ.get('DB_USER', 'fbsr_pontun')
host = os.environ.get('DB_HOST', '127.0.0.1')
password = os.environ.get('DB_PASSWORD', '')

# Use the environment variables in the connection string
conn = psycopg2.connect(f"dbname={dbname} user={user} host={host} password={password}")

cur  = conn.cursor()

def main(listing_id, filename):
    saved, skipped, added = 0, 0, set()

    workbook = openpyxl.load_workbook(filename, data_only=True)
    sheets = {
        None:         workbook['Búnaðar verðlisti'], 
        'Aku':        workbook['Aku verðlisti'], 
        'Icebreaker': workbook['Icebreaker verðlisti'], 
        'Bliz':       workbook['Bliz verðlisti'], 
    }

    for sheet_idx, (sheet_name, worksheet) in enumerate(sheets.items()):
        print(20*"=", sheet_name)
        idx = 1

        category = None
        for row in range(8,worksheet.max_row+1):
            vendor_id = (sheet_idx+1)*10000 + idx
            name      = worksheet.cell(column=2, row=row).value
            try:
                url       = worksheet.cell(column=2, row=row).hyperlink.target
            except AttributeError:
                url = None
            desc      = worksheet.cell(column=3, row=row).value
            try:
                price     = float(worksheet.cell(column=5, row=row).value)
            except (TypeError, ValueError):
                price     = None
            desc      = "" if pd.isnull(desc) else desc

            if not pd.isnull(price):
                if sheet_name is not None:
                    if sheet_name.startswith('Ice'):
                        name = f'{sheet_name} {name} ({worksheet.cell(column=1, row=row).value})'
                    else:
                        name = f'{sheet_name} {name}'
                cur.execute(f"INSERT INTO items (listing_id, item_name, vendor_id, price, description, vendor_url) VALUES (%s, %s, %s, %s, %s, %s) RETURNING id", (listing_id, name, vendor_id, price, desc, url))
                item_id = cur.fetchone()[0]
                if category is not None:
                    cur.execute(f"INSERT INTO item_properties (item_id, original, adjusted, value) VALUES (%s, %s, %s, %s)", (item_id, "Flokkur", "Flokkur", category)) 
                print(f"Inserting {category} : {name} : {vendor_id} : {price} : {desc} : {url}")
                saved += 1
                idx += 1
            else:
                if sheet_name in ("Aku", None) and (not pd.isnull(worksheet.cell(column=1,row=row).value)):
                    category = worksheet.cell(column=1,row=row).value
                if not (pd.isnull(worksheet.cell(column=1,row=row).value) and pd.isnull(name)):
                    print(f"Skipped row item {worksheet.cell(column=1,row=row).value} : {name} : {url} : {sheet_name}")
                skipped += 1
    print(f"Saved {saved} skipped {skipped}")

    conn.commit()
    import pdb; pdb.set_trace()
    #        cur.execute(f"INSERT INTO items (listing_id, item_name, vendor_id, vendor_url, price) VALUES ({listing_id}, %s, %s, %s, %s) RETURNING id", (i['title'], i['sku'], i['url'], i['price']))
    #        item_id = cur.fetchone()

    #        for prop, val in i['props']:
    #            cur.execute("INSERT INTO item_properties (item_id, original, adjusted, value) VALUES (%s, %s, %s, %s)", (item_id, prop, prop, val))
    #        conn.commit()
    #        saved += 1
    #    except Exception as e:
    #        print(e)
    #        import pdb; pdb.set_trace()
    #print(f"Saved {saved} items to database")



if __name__ == "__main__":
    assert len(sys.argv) > 2, "Missing listing id or filename"
    main(int(sys.argv[1]), sys.argv[2])
