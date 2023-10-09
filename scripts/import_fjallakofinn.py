    import sys
    import pandas as pd
    import itertools
    import openpyxl

    import psycopg2

    # Fetch environment variables
    dbname = os.environ.get('DB_NAME', 'fbsr_pontun')
    user = os.environ.get('DB_USER', 'fbsr_pontun')
    host = os.environ.get('DB_HOST', '127.0.0.1')
    password = os.environ.get('DB_PASSWORD', '')

    # Use the environment variables in the connection string
    conn = psycopg2.connect(f"dbname={dbname} user={user} host={host} password={password}")
    cur  = conn.cursor()

    def main(listing_id, filename):
        workbook = openpyxl.load_workbook(filename, data_only=True)
        worksheet = workbook.active
        saved, skipped, added = 0, 0, set()
        category = None
        for row in range(1,worksheet.max_row+1):
            _row = []
            vendor_id = worksheet.cell(column=1, row=row).value
            name      = worksheet.cell(column=2, row=row).value
            try:
                url       = worksheet.cell(column=2, row=row).hyperlink.target
            except AttributeError:
                url = None
            desc      = worksheet.cell(column=3, row=row).value
            try:
                price     = float(worksheet.cell(column=5, row=row).value)
            except (TypeError, ValueError):
            price     = None
        desc      = "" if pd.isnull(desc) else desc

        if not pd.isnull(price):
            cur.execute(f"INSERT INTO items (listing_id, item_name, vendor_id, price, description, vendor_url) VALUES (%s, %s, %s, %s, %s, %s) RETURNING id", (listing_id, name, vendor_id, price, desc, url))
            item_id = cur.fetchone()[0]
            if category is not None:
                cur.execute(f"INSERT INTO item_properties (item_id, original, adjusted, value) VALUES (%s, %s, %s, %s)", (item_id, "Flokkur", "Flokkur", category)) 
            #print(f"{listing_id} : {name} : {vendor_id} : {price} : {desc} : {url}")
            saved += 1
        else:
            if not pd.isnull(vendor_id):
                print(f"Skipped row item {vendor_id} : {name}")
                category = vendor_id
                if category.startswith("Fjallahjól"):
                    category = "Fjallahjól"
            skipped += 1
    print(f"Saved {saved} skipped {skipped} pandas df has {df.loc[~pd.isnull(pd.to_numeric(df.iloc[:,4], errors='coerce'))].shape[0]} rows")

    conn.commit()
    import pdb; pdb.set_trace()
    #        cur.execute(f"INSERT INTO items (listing_id, item_name, vendor_id, vendor_url, price) VALUES ({listing_id}, %s, %s, %s, %s) RETURNING id", (i['title'], i['sku'], i['url'], i['price']))
    #        item_id = cur.fetchone()

    #        for prop, val in i['props']:
    #            cur.execute("INSERT INTO item_properties (item_id, original, adjusted, value) VALUES (%s, %s, %s, %s)", (item_id, prop, prop, val))
    #        conn.commit()
    #        saved += 1
    #    except Exception as e:
    #        print(e)
    #        import pdb; pdb.set_trace()
    #print(f"Saved {saved} items to database")



if __name__ == "__main__":
    assert len(sys.argv) > 2, "Missing listing id or filename"
    main(int(sys.argv[1]), sys.argv[2])
